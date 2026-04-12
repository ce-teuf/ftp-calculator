"""
Génération de matrices de taux fictives pour le Module 1 du FTP Simulator.

Chaque matrice suit le format attendu par le backend :
  - Colonne A : date_month (YYYY-MM)
  - Colonne B : period_type (observed | contrafactual | projected)
  - Colonnes C+ : valeurs par tenor (1M, 3M, 6M, 12M, 24M, 60M, 120M)

Utilisation :
    python generate_rate_matrices.py          # génère les fichiers xlsx
    python generate_rate_matrices.py --seed   # génère + envoie à l'API

Dépendances : openpyxl, requests
    pip install openpyxl requests
"""

import argparse
import math
import os
import random
import requests
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Configuration ──────────────────────────────────────────────────────────────

OUT_DIR   = Path(__file__).parent / "output"
API_URL   = "http://localhost:3000/api/rate-matrices"
TENORS    = ["1M", "3M", "6M", "12M", "24M", "60M", "120M"]

# Plage historique : 2022-01 → 2025-03 (observed)
# Projection       : 2025-04 → 2025-12 (projected)
OBS_START   = (2022, 1)
OBS_END     = (2025, 3)
PROJ_START  = (2025, 4)
PROJ_END    = (2025, 12)


def date_range(start: tuple, end: tuple):
    """Génère une liste de (year, month) de start à end inclus."""
    y, m = start
    ey, em = end
    while (y, m) <= (ey, em):
        yield (y, m)
        m += 1
        if m > 12:
            m = 1
            y += 1


def months_since(ref_year: int, ref_month: int, year: int, month: int) -> int:
    return (year - ref_year) * 12 + (month - ref_month)


# ── Fonctions de forme de courbe ───────────────────────────────────────────────

def tenor_months(tenor: str) -> float:
    """Convertit un label de tenor en mois."""
    mapping = {"1M": 1, "3M": 3, "6M": 6, "12M": 12, "24M": 24, "60M": 60, "120M": 120}
    return float(mapping[tenor])


def nelson_siegel(beta0, beta1, beta2, lam, maturity_months):
    """Courbe Nelson-Siegel standard (maturity en mois)."""
    m = maturity_months
    l = lam
    if m == 0:
        return beta0
    x = m / l
    ex = math.exp(-x)
    return (
        beta0
        + beta1 * (1 - ex) / x
        + beta2 * ((1 - ex) / x - ex)
    )


# ── Définitions des matrices ───────────────────────────────────────────────────

MATRICES = [
    {
        "name":         "Base Rate EUR (ESTR)",
        "description":  "Courbe de taux sans risque EUR — ESTR / OIS",
        "currency":     "EUR",
        "risks":        ["base_rate"],
        "interp_method": "cubic",
        # β0 évolue de -0.5% (2022-01) à 4.0% (2023-09) puis redescend à 2.5% (2025-03)
        # β1 : légère inversion initiale puis normale
        # Projection : glissement progressif vers 2.0%
        "curve_fn": lambda t: {  # t = mois depuis 2022-01
            "beta0": (
                -0.005 + min(t, 20) * 0.002  # montée jusqu'à t=20 (sep 2023)
                - max(0, t - 20) * 0.0012    # descente ensuite
            ),
            "beta1": -0.002 + t * 0.0001,
            "beta2": 0.003,
            "lam": 24.0,
        },
    },
    {
        "name":         "Credit Spread EUR",
        "description":  "Z-spread senior unsecured, courbe crédit EUR",
        "currency":     "EUR",
        "risks":        ["credit_spread"],
        "interp_method": "linear",
        # Spread de crédit stable autour de 0.6-0.9% selon le tenor
        "curve_fn": lambda t: {
            "beta0": 0.007 + 0.0002 * math.sin(t * 0.3),
            "beta1": -0.002,
            "beta2": 0.001,
            "lam": 36.0,
        },
    },
    {
        "name":         "Term Liquidity Premium EUR",
        "description":  "Prime de liquidité à terme — LCR/NSFR",
        "currency":     "EUR",
        "risks":        ["tlp"],
        "interp_method": "linear",
        "curve_fn": lambda t: {
            "beta0": 0.002 + max(0, 15 - t) * 0.0001,  # légère décroissance historique
            "beta1": 0.001,
            "beta2": 0.0005,
            "lam": 48.0,
        },
    },
    {
        "name":         "Charge Liquidité Réglementaire EUR",
        "description":  "Coussin LCR/NSFR — coût réglementaire de la liquidité",
        "currency":     "EUR",
        "risks":        ["clp"],
        "interp_method": "flat_forward",
        "curve_fn": lambda t: {
            "beta0": 0.0015,
            "beta1": 0.0005,
            "beta2": 0.0,
            "lam": 60.0,
        },
    },
    {
        "name":         "Base Rate USD (SOFR)",
        "description":  "Courbe de taux sans risque USD — SOFR OIS",
        "currency":     "USD",
        "risks":        ["base_rate"],
        "interp_method": "cubic",
        "curve_fn": lambda t: {
            "beta0": (
                0.002 + min(t, 16) * 0.003
                - max(0, t - 16) * 0.0015
            ),
            "beta1": -0.003 + t * 0.00015,
            "beta2": 0.004,
            "lam": 20.0,
        },
    },
    {
        "name":         "FTP Spread Global (base_rate + credit_spread)",
        "description":  "Taux de base + spread crédit — composantes indissociables",
        "currency":     "EUR",
        "risks":        ["base_rate", "credit_spread"],  # 2 risques → indissociables
        "interp_method": "cubic",
        "curve_fn": lambda t: {
            "beta0": (
                0.002 + min(t, 20) * 0.002
                - max(0, t - 20) * 0.0012
                + 0.007
            ),
            "beta1": -0.003,
            "beta2": 0.004,
            "lam": 24.0,
        },
    },
]


# ── Génération des données ────────────────────────────────────────────────────

def generate_matrix_data(matrix_def: dict) -> list[dict]:
    """Génère les lignes d'une matrice sous forme de dicts."""
    rows = []
    noise = random.Random(hash(matrix_def["name"]))  # reproductible

    for (y, m) in date_range(OBS_START, OBS_END):
        t   = months_since(*OBS_START, y, m)
        cfg = matrix_def["curve_fn"](t)
        row = {
            "date_month":  f"{y:04d}-{m:02d}",
            "period_type": "observed",
        }
        for tenor in TENORS:
            val = nelson_siegel(
                cfg["beta0"], cfg["beta1"], cfg["beta2"],
                cfg["lam"], tenor_months(tenor)
            )
            # Petit bruit aléatoire ±1bp
            val += noise.uniform(-0.0001, 0.0001)
            row[tenor] = max(round(val, 6), -0.02)  # pas en dessous de -2%
        rows.append(row)

    for (y, m) in date_range(PROJ_START, PROJ_END):
        t   = months_since(*OBS_START, y, m)
        cfg = matrix_def["curve_fn"](t)
        row = {
            "date_month":  f"{y:04d}-{m:02d}",
            "period_type": "projected",
        }
        for tenor in TENORS:
            val = nelson_siegel(
                cfg["beta0"], cfg["beta1"], cfg["beta2"],
                cfg["lam"], tenor_months(tenor)
            )
            row[tenor] = max(round(val, 6), -0.02)
        rows.append(row)

    return rows


# ── Écriture Excel ────────────────────────────────────────────────────────────

HEADER_FILL_OBS  = PatternFill("solid", fgColor="DBEAFE")   # bleu clair
HEADER_FILL_PROJ = PatternFill("solid", fgColor="FEF3C7")   # jaune clair
HEADER_FONT      = Font(bold=True)


def write_xlsx(matrix_def: dict, rows: list[dict], out_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Taux"

    # Entêtes
    cols = ["date_month", "period_type"] + TENORS
    for col_idx, header in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.fill      = HEADER_FILL_OBS

    # Données
    for row_idx, row in enumerate(rows, start=2):
        fill = HEADER_FILL_PROJ if row["period_type"] == "projected" else None
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            if col_idx > 2:
                cell.number_format = "0.0000%"
            if fill:
                cell.fill = fill

    # Largeur des colonnes
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    for col_letter in "CDEFGHI":
        ws.column_dimensions[col_letter].width = 10

    wb.save(out_path)
    print(f"  Écrit : {out_path.name}")


# ── Seed via API ───────────────────────────────────────────────────────────────

def seed_via_api(matrix_def: dict, xlsx_path: Path):
    """Envoie une matrice à l'API /api/rate-matrices."""
    with open(xlsx_path, "rb") as f:
        files   = {"file": (xlsx_path.name, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data    = {
            "name":          matrix_def["name"],
            "description":   matrix_def.get("description", ""),
            "currency":      matrix_def["currency"],
            "interp_method": matrix_def["interp_method"],
            "status":        "active",
        }
        # Les risk_keys en multipart repeated fields
        for rk in matrix_def["risks"]:
            data[f"risk_key"] = rk  # note: requests ne supporte pas les listes comme ça

    # requests supporte les tuples pour champs répétés
    with open(xlsx_path, "rb") as f:
        files = [("file", (xlsx_path.name, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))]
        form_fields = [
            ("name",          matrix_def["name"]),
            ("description",   matrix_def.get("description", "")),
            ("currency",      matrix_def["currency"]),
            ("interp_method", matrix_def["interp_method"]),
            ("status",        "active"),
        ]
        for rk in matrix_def["risks"]:
            form_fields.append(("risk_key", rk))

        resp = requests.post(API_URL, data=form_fields, files=files, timeout=15)

    if resp.status_code in (200, 201):
        created = resp.json()
        print(f"  Seedé  : {matrix_def['name']} → id={created['id']}")
    else:
        print(f"  ERREUR {resp.status_code} pour '{matrix_def['name']}' : {resp.text[:200]}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Génération de matrices de taux fictives")
    parser.add_argument("--seed", action="store_true", help="Envoie les fichiers à l'API backend")
    args = parser.parse_args()

    OUT_DIR.mkdir(exist_ok=True)
    random.seed(42)

    print(f"Génération de {len(MATRICES)} matrices de taux...")
    for matrix_def in MATRICES:
        print(f"\n→ {matrix_def['name']} ({matrix_def['currency']}, risques: {matrix_def['risks']})")

        rows      = generate_matrix_data(matrix_def)
        safe_name = matrix_def["name"].replace(" ", "_").replace("/", "-")[:40]
        out_path  = OUT_DIR / f"{safe_name}.xlsx"

        write_xlsx(matrix_def, rows, out_path)

        if args.seed:
            seed_via_api(matrix_def, out_path)

    print(f"\nTerminé. Fichiers dans : {OUT_DIR}")
    if args.seed:
        print("Matrices envoyées à l'API.")
    else:
        print("Pour envoyer à l'API : python generate_rate_matrices.py --seed")


if __name__ == "__main__":
    main()
