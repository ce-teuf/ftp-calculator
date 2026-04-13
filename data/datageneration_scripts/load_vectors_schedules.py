#!/usr/bin/env python3
"""
load_vectors_schedules.py
Convertit les CSV générés par les scripts existants en XLSX compatibles
avec le backend FTP Simulator, puis les charge via l'API.

Format attendu par le backend :
  Vectors  : date_month (YYYY-MM) | period_type | value
  Schedules: date_month (YYYY-MM) | period_type | 1M | 3M | 6M | 12M | 24M | 60M | 120M

Usage :
    python load_vectors_schedules.py
    python load_vectors_schedules.py --dry-run   # génère les XLSX sans uploader
"""

import argparse
import csv
import io
import os
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests

# ── Config ─────────────────────────────────────────────────────────────────────

API         = "http://localhost:3000/api"
SCRIPTS_DIR = Path(__file__).parent

# Dates pivot : avant cette date → observed, à partir de → projected
# Aligner sur la même convention que les matrices (observed jusqu'à 2024-12)
PROJ_FROM = (2025, 1)

# ── Styles XLSX ────────────────────────────────────────────────────────────────

BOLD      = Font(bold=True)
HDR_FILL  = PatternFill("solid", fgColor="DBEAFE")
PROJ_FILL = PatternFill("solid", fgColor="FEF9C3")
CENTER    = Alignment(horizontal="center")


def _xlsx_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _style_header(ws, n_cols: int):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.font = BOLD
        c.fill = HDR_FILL
        c.alignment = CENTER


# ── Conversion de date MM-YYYY → YYYY-MM ──────────────────────────────────────

def convert_date(raw: str) -> tuple[str, str]:
    """
    Accepte MM-YYYY ou YYYY-MM.
    Retourne (date_ym: str "YYYY-MM", period_type: str).
    """
    raw = raw.strip()
    if len(raw) == 7 and raw[2] == '-':          # MM-YYYY
        mm, yyyy = raw[:2], raw[3:]
    elif len(raw) == 7 and raw[4] == '-':        # YYYY-MM
        yyyy, mm = raw[:4], raw[5:]
    else:
        raise ValueError(f"Date non reconnue : '{raw}'")
    year, month = int(yyyy), int(mm)
    date_ym = f"{year:04d}-{month:02d}"
    pt = "observed" if (year, month) < PROJ_FROM else "projected"
    return date_ym, pt


# ── Conversion Vector CSV → XLSX ──────────────────────────────────────────────

def vector_csv_to_xlsx(csv_path: Path) -> bytes:
    """
    Lit un CSV avec colonnes [date, outstanding] ou [date, value].
    Produit un XLSX avec [date_month, period_type, value].
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outstanding"

    headers = ["date_month", "period_type", "value"]
    ws.append(headers)
    _style_header(ws, len(headers))

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        value_col = None
        for row in reader:
            if value_col is None:
                # Trouver la colonne de valeur (outstanding ou value)
                value_col = next(
                    (k for k in row if k.lower() in ("outstanding", "value")),
                    None
                )
                if value_col is None:
                    raise ValueError(f"Colonne 'outstanding' ou 'value' introuvable dans {csv_path.name}")

            date_ym, pt = convert_date(row["date"])
            val = float(row[value_col])
            data_row = [date_ym, pt, round(val, 4)]
            ws.append(data_row)
            if pt == "projected":
                for col in range(1, 4):
                    ws.cell(row=ws.max_row, column=col).fill = PROJ_FILL

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    return _xlsx_bytes(wb)


# ── Conversion Schedule CSV → XLSX ────────────────────────────────────────────

def schedule_csv_to_xlsx(csv_path: Path) -> bytes:
    """
    Lit un CSV avec colonnes [date, 1, 2, 3, …, N] (mois numérotés).
    Passe TOUTES les colonnes mensuelles dans le XLSX sans filtrage.
    Produit un XLSX avec [date_month, period_type, 1, 2, 3, …, N].
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        # Toutes les colonnes numériques = mois, triées
        month_cols = sorted(int(k) for k in fieldnames if k.isdigit())
        if not month_cols:
            raise ValueError(f"Aucune colonne mensuelle trouvée dans {csv_path.name}")
        month_labels = [str(m) for m in month_cols]

        headers = ["date_month", "period_type"] + month_labels
        ws.append(headers)
        _style_header(ws, len(headers))

        for row in reader:
            date_ym, pt = convert_date(row["date"])
            buckets = [round(float(row.get(str(m), 0) or 0), 6) for m in month_cols]
            ws.append([date_ym, pt] + buckets)
            if pt == "projected":
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = PROJ_FILL

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    for i in range(len(month_labels)):
        col_letter = openpyxl.utils.get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = max(4, len(month_labels[i]) + 1)
    return _xlsx_bytes(wb)


# ── Upload helpers ─────────────────────────────────────────────────────────────

def upload_vector(name: str, desc: str, xlsx: bytes, dry_run: bool) -> str | None:
    if dry_run:
        out = SCRIPTS_DIR / "data_vector_outs" / f"{name.replace(' ', '_')}.xlsx"
        out.write_bytes(xlsx)
        print(f"  [dry-run] Écrit : {out}")
        return None
    files = [("file", (f"{name}.xlsx", xlsx,
              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))]
    data  = [("name", name), ("description", desc)]
    r = requests.post(f"{API}/outstanding-vectors", data=data, files=files, timeout=30)
    if not r.ok:
        print(f"  ✗ {name}: {r.status_code} — {r.text[:200]}")
        return None
    vid = r.json()["id"]
    print(f"  ✓ {name:<50} → {vid}")
    return vid


def upload_schedule(name: str, desc: str, schedule_type: str, xlsx: bytes, dry_run: bool) -> str | None:
    if dry_run:
        out = SCRIPTS_DIR / "data_schedules" / f"{name.replace(' ', '_')}.xlsx"
        out.write_bytes(xlsx)
        print(f"  [dry-run] Écrit : {out}  [{schedule_type}]")
        return None
    files = [("file", (f"{name}.xlsx", xlsx,
              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))]
    data  = [("name", name), ("description", desc), ("schedule_type", schedule_type)]
    r = requests.post(f"{API}/amort-schedules", data=data, files=files, timeout=30)
    if not r.ok:
        print(f"  ✗ {name}: {r.status_code} — {r.text[:200]}")
        return None
    sid = r.json()["id"]
    print(f"  ✓ {name:<50} [{schedule_type}] → {sid}")
    return sid


# ── Définitions ────────────────────────────────────────────────────────────────

VECTOR_DEFS = [
    {
        "csv":  "data_vector_outs/convex_up.csv",
        "name": "Convex Growth 2020-2024",
        "desc": "Encours à croissance convexe (1k → 8k) avec bruit absolu — généré depuis convex_trend_mild_noise.yaml",
    },
    {
        "csv":  "data_vector_outs/logistic_growth.csv",
        "name": "Logistic Growth 2020-2024",
        "desc": "Encours en courbe S logistique (500 → 5k) avec bruit relatif — généré depuis logistic_trend.yaml",
    },
    {
        "csv":  "data_vector_outs/exponential_down.csv",
        "name": "Exponential Decline 2020-2024",
        "desc": "Encours en décroissance exponentielle (10k → 500) avec bruit relatif 5% — généré depuis trend_exp_down.yaml",
    },
    {
        "csv":  "data_vector_outs/linear_up.csv",
        "name": "Linear Growth 2020-2024",
        "desc": "Encours en croissance linéaire (1k → 5k) avec bruit absolu — généré depuis trend_lineair_abs_noise.yaml",
    },
    {
        "csv":  "data_vector_outs/concave_uniform.csv",
        "name": "Concave Growth 2021-2022",
        "desc": "Encours à croissance concave — série courte (30 mois)",
    },
    {
        "csv":  "data_vector_outs/convex_normal.csv",
        "name": "Convex Stable 2020-2024",
        "desc": "Encours convexe stable (≈260) avec bruit normal — série de référence",
    },
]

SCHEDULE_DEFS = [
    # ── Stock schedules (book existant) ──────────────────────────────────────
    {
        "csv":   "data_schedules/resultats_matrix.csv",
        "name":  "Schedule 5Y — Stock",
        "desc":  "Stock : amortissement 3 périodes : 60M → 30M (2021) → 12M (2023)",
        "stype": "stock",
    },
    {
        "csv":   "data_schedules/schedule_24m.csv",
        "name":  "Schedule 2Y — Stock",
        "desc":  "Stock : amortissement linéaire 24 mois (curvature 1.0, std 0.2) — 2022-2025",
        "stype": "stock",
    },
    {
        "csv":   "data_schedules/schedule_120m.csv",
        "name":  "Schedule 10Y — Stock",
        "desc":  "Stock : amortissement 10 ans (120 mois, curvature 1.5, std 0.1) — 2022-2025",
        "stype": "stock",
    },
    {
        "csv":   "data_schedules/schedule_240m.csv",
        "name":  "Schedule 20Y — Stock",
        "desc":  "Stock : amortissement 20 ans (240 mois, curvature 1.5, std 0.15) — 2022-2025",
        "stype": "stock",
    },
    # ── Nouvelle production schedules (flux entrants) ─────────────────────────
    {
        "csv":   "data_schedules/schedule_24m.csv",
        "name":  "Schedule 2Y — Nvl. Production",
        "desc":  "Nvl. prod. : profil de nouvelle production 24 mois — 2022-2025",
        "stype": "new_production",
    },
    {
        "csv":   "data_schedules/schedule_120m.csv",
        "name":  "Schedule 10Y — Nvl. Production",
        "desc":  "Nvl. prod. : profil de nouvelle production 10 ans (120 mois) — 2022-2025",
        "stype": "new_production",
    },
    {
        "csv":   "data_schedules/schedule_240m.csv",
        "name":  "Schedule 20Y — Nvl. Production",
        "desc":  "Nvl. prod. : profil de nouvelle production 20 ans (240 mois) — 2022-2025",
        "stype": "new_production",
    },
]


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Load vectors & schedules into FTP Simulator")
    parser.add_argument("--dry-run", action="store_true",
                        help="Produit les XLSX sans les envoyer à l'API")
    args = parser.parse_args()

    if not args.dry_run:
        try:
            requests.get(f"{API}/risk-types", timeout=5).raise_for_status()
            print(f"Backend OK — {API}\n")
        except Exception as e:
            print(f"✗ Backend inaccessible : {e}")
            sys.exit(1)

    # ── Vectors ──
    print("=" * 60)
    print("  Outstanding Vectors")
    print("=" * 60)
    v_ok = v_err = 0
    for vdef in VECTOR_DEFS:
        csv_path = SCRIPTS_DIR / vdef["csv"]
        if not csv_path.exists():
            print(f"  ⚠ CSV introuvable : {csv_path} — skip")
            v_err += 1
            continue
        try:
            xlsx = vector_csv_to_xlsx(csv_path)
            result = upload_vector(vdef["name"], vdef["desc"], xlsx, args.dry_run)
            if result or args.dry_run:
                v_ok += 1
            else:
                v_err += 1
        except Exception as e:
            print(f"  ✗ {vdef['name']}: {e}")
            v_err += 1

    # ── Schedules ──
    print()
    print("=" * 60)
    print("  Amortization Schedules")
    print("=" * 60)
    s_ok = s_err = 0
    for sdef in SCHEDULE_DEFS:
        csv_path = SCRIPTS_DIR / sdef["csv"]
        if not csv_path.exists():
            print(f"  ⚠ CSV introuvable : {csv_path} — skip")
            s_err += 1
            continue
        try:
            xlsx = schedule_csv_to_xlsx(csv_path)
            result = upload_schedule(sdef["name"], sdef["desc"], sdef.get("stype", "stock"), xlsx, args.dry_run)
            if result or args.dry_run:
                s_ok += 1
            else:
                s_err += 1
        except Exception as e:
            print(f"  ✗ {sdef['name']}: {e}")
            s_err += 1

    print()
    print("=" * 60)
    print(f"  Vectors  : {v_ok} chargés, {v_err} erreurs")
    print(f"  Schedules: {s_ok} chargés, {s_err} erreurs")
    print("=" * 60)


if __name__ == "__main__":
    main()
