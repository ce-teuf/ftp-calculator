#!/usr/bin/env python3
"""
seed_api.py — Seed complet du FTP Simulator via l'API REST.

Ordre d'exécution :
  1. Rate matrices  (via generate_rate_matrices.py --seed)
  2. Outstanding vectors   (XLSX → POST /api/outstanding-vectors)
  3. Amortization schedules (XLSX → POST /api/amort-schedules)
  4. Portfolios + pairs    (JSON → POST /api/portfolios/...)
  5. Hypercubes            (JSON → POST /api/hypercubes)
  6. Study units + assignments  (JSON → POST /api/study-units/...)
  7. Validate study units  (POST /api/study-units/:id/validate)
  8. Studies               (JSON → POST /api/studies/...)

Usage :
    python seed_api.py              # seed tout
    python seed_api.py --skip-matrices  # skip l'étape matrices (déjà faites)
"""

from __future__ import annotations

import argparse
import io
import json
import math
import os
import random
import subprocess
import sys
from typing import NamedTuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests

# ── Config ─────────────────────────────────────────────────────────────────────

API = "http://localhost:3000/api"
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))

# Tenors utilisés dans les matrices et les schedules
TENORS       = ["1M", "3M", "6M", "12M", "24M", "60M", "120M"]
TENOR_MONTHS = [1,    3,    6,    12,    24,    60,    120]

# Plages temporelles
OBS_START  = (2022, 1)
OBS_END    = (2024, 12)
PROJ_START = (2025, 1)
PROJ_END   = (2025, 12)

random.seed(42)


# ── Utilitaires dates ──────────────────────────────────────────────────────────

def date_range(start: tuple[int,int], end: tuple[int,int]):
    y, m = start
    ey, em = end
    while (y, m) <= (ey, em):
        yield y, m
        m += 1
        if m > 12:
            m = 1; y += 1


def all_dates() -> list[tuple[str, str]]:
    """Retourne [(date_ym, period_type)] de 2022-01 à 2025-12."""
    rows: list[tuple[str, str]] = []
    for y, m in date_range(OBS_START, OBS_END):
        rows.append((f"{y:04d}-{m:02d}", "observed"))
    for y, m in date_range(PROJ_START, PROJ_END):
        rows.append((f"{y:04d}-{m:02d}", "projected"))
    return rows


# ── Utilitaires XLSX ───────────────────────────────────────────────────────────

BOLD      = Font(bold=True)
HDR_FILL  = PatternFill("solid", fgColor="DBEAFE")
PROJ_FILL = PatternFill("solid", fgColor="FEF9C3")
CENTER    = Alignment(horizontal="center")


def _xlsx_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_vector_xlsx(rows_data: list[tuple[str, str, float]]) -> bytes:
    """rows_data = [(date_ym, period_type, value), ...]"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outstanding"
    headers = ["date_month", "period_type", "value"]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.font = BOLD; c.fill = HDR_FILL; c.alignment = CENTER
    for date, pt, val in rows_data:
        row = [date, pt, round(val, 2)]
        ws.append(row)
        if pt == "projected":
            for col in range(1, 4):
                ws.cell(row=ws.max_row, column=col).fill = PROJ_FILL
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    return _xlsx_bytes(wb)


def make_schedule_xlsx(rows_data: list[tuple[str, str, list[float]]], bucket_labels: list[str]) -> bytes:
    """rows_data = [(date_ym, period_type, [bucket values]), ...]"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    headers = ["date_month", "period_type"] + bucket_labels
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.font = BOLD; c.fill = HDR_FILL; c.alignment = CENTER
    for date, pt, buckets in rows_data:
        ws.append([date, pt] + [round(b, 6) for b in buckets])
        if pt == "projected":
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = PROJ_FILL
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    for i in range(len(bucket_labels)):
        col = chr(ord("C") + i) if i < 20 else None
        if col:
            ws.column_dimensions[col].width = 10
    return _xlsx_bytes(wb)


# ── Utilitaires profils d'amortissement ───────────────────────────────────────

def amort_at_tenors(maturity_months: int, curvature: float,
                    tenor_months: list[int] = TENOR_MONTHS) -> list[float]:
    """Remaining fraction at each tenor month — used as schedule bucket values."""
    if maturity_months <= 1:
        return [0.0] * len(tenor_months)
    denom = maturity_months - 1
    result = []
    for t in tenor_months:
        if t >= maturity_months:
            result.append(0.0)
        else:
            ratio = (1.0 - (t - 1) / denom) ** curvature
            result.append(round(max(0.0, ratio), 6))
    return result


def nmd_at_tenors(lam: float, core_ratio: float,
                  tenor_months: list[int] = TENOR_MONTHS) -> list[float]:
    """Exponential-decay NMD profile at tenor buckets."""
    result = []
    for t in tenor_months:
        val = core_ratio + (1.0 - core_ratio) * math.exp(-lam * t / 12.0)
        result.append(round(max(0.0, val), 6))
    return result


# ── Utilitaires trends outstandings ───────────────────────────────────────────

def trend_linear(n, start, end, noise_std=0.0) -> list[float]:
    vals = [start + (end - start) * i / max(n - 1, 1) for i in range(n)]
    return [max(0.0, v + random.gauss(0, noise_std)) for v in vals]

def trend_exponential(n, start, end, noise_std=0.0) -> list[float]:
    if start <= 0 or end <= 0:
        return trend_linear(n, start, end, noise_std)
    vals = [start * (end / start) ** (i / max(n - 1, 1)) for i in range(n)]
    return [max(0.0, v + random.gauss(0, noise_std)) for v in vals]

def trend_logistic(n, start, end, midpoint=0.5, steepness=10.0, noise_std=0.0) -> list[float]:
    def logit(x): return 1.0 / (1.0 + math.exp(-steepness * (x - midpoint)))
    vals = [start + (end - start) * logit(i / max(n - 1, 1)) for i in range(n)]
    return [max(0.0, v + random.gauss(0, noise_std)) for v in vals]

def trend_nmd(n, base, lam, noise_frac=0.0) -> list[float]:
    vals = [base * math.exp(-lam * i / 12.0) for i in range(n)]
    return [max(0.0, v + random.gauss(0, noise_frac * base)) for v in vals]


# ── HTTP helpers ───────────────────────────────────────────────────────────────

def post_json(path: str, data: dict) -> dict:
    r = requests.post(f"{API}{path}", json=data, timeout=30)
    if not r.ok:
        raise RuntimeError(f"POST {path} → {r.status_code}: {r.text[:300]}")
    return r.json()

def put_json(path: str, data: dict) -> dict:
    r = requests.put(f"{API}{path}", json=data, timeout=30)
    if not r.ok:
        raise RuntimeError(f"PUT {path} → {r.status_code}: {r.text[:300]}")
    return r.json()

def get_json(path: str) -> dict | list:
    r = requests.get(f"{API}{path}", timeout=30)
    if not r.ok:
        raise RuntimeError(f"GET {path} → {r.status_code}: {r.text[:300]}")
    return r.json()

def post_file(path: str, name: str, desc: str, xlsx_bytes: bytes,
              extra_fields: list[tuple] | None = None) -> dict:
    files = [("file", (f"{name}.xlsx", xlsx_bytes,
              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))]
    data  = [("name", name), ("description", desc)]
    if extra_fields:
        data.extend(extra_fields)
    r = requests.post(f"{API}{path}", data=data, files=files, timeout=30)
    if not r.ok:
        raise RuntimeError(f"POST {path} → {r.status_code}: {r.text[:300]}")
    return r.json()

def section(title: str):
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")

def ok(label: str, id_: str):
    print(f"  ✓ {label:<45} id={id_}")


# ── Étape 1 : Rate matrices ────────────────────────────────────────────────────

def seed_matrices():
    section("1 — Rate matrices")
    script = os.path.join(SCRIPTS_DIR, "data_rate_matrices", "generate_rate_matrices.py")
    result = subprocess.run(
        [sys.executable, script, "--seed"],
        capture_output=True, text=True, cwd=SCRIPTS_DIR
    )
    if result.returncode != 0:
        print("  STDERR:", result.stderr[:500])
        raise RuntimeError("generate_rate_matrices.py a échoué")
    for line in result.stdout.strip().split("\n"):
        if line.strip():
            print(f"  {line}")


# ── Étape 2 : Outstanding vectors ─────────────────────────────────────────────

class VectorSpec(NamedTuple):
    name: str
    desc: str
    trend_fn: str
    params: dict

VECTOR_SPECS: list[VectorSpec] = [
    VectorSpec(
        "Retail Mortgages 20Y",
        "Portefeuille hypothécaire résidentiel taux fixe 20 ans — croissance linéaire",
        "linear",
        {"start": 450e6, "end": 720e6, "noise": 8e6},
    ),
    VectorSpec(
        "SME Loans 5Y",
        "Prêts PME 5 ans — croissance exponentielle",
        "exponential",
        {"start": 80e6, "end": 190e6, "noise": 2e6},
    ),
    VectorSpec(
        "NMD Demand Deposits Core",
        "Dépôts à vue retail — comportemental décroissance exponentielle (core 75%)",
        "nmd",
        {"base": 2_000e6, "lam": 0.05, "noise": 0.005},
    ),
    VectorSpec(
        "Corporate Loans 7Y",
        "Prêts entreprises 7 ans — book stable",
        "linear",
        {"start": 250e6, "end": 260e6, "noise": 6e6},
    ),
    VectorSpec(
        "Consumer Credit 2Y",
        "Crédits consommation 2 ans — forte croissance",
        "exponential",
        {"start": 50e6, "end": 130e6, "noise": 2e6},
    ),
    VectorSpec(
        "Commercial Real Estate 15Y",
        "Immobilier commercial 15 ans — Paris CBD, croissance modérée",
        "linear",
        {"start": 600e6, "end": 750e6, "noise": 10e6},
    ),
]


def seed_vectors() -> dict[str, str]:
    """Returns {name → id}"""
    section("2 — Outstanding vectors")
    dates = all_dates()
    n = len(dates)
    ids: dict[str, str] = {}

    for spec in VECTOR_SPECS:
        p = spec.params
        if spec.trend_fn == "linear":
            vals = trend_linear(n, p["start"], p["end"], p.get("noise", 0))
        elif spec.trend_fn == "exponential":
            vals = trend_exponential(n, p["start"], p["end"], p.get("noise", 0))
        elif spec.trend_fn == "logistic":
            vals = trend_logistic(n, p["start"], p["end"],
                                  p.get("midpoint", 0.5), p.get("steepness", 10), p.get("noise", 0))
        elif spec.trend_fn == "nmd":
            vals = trend_nmd(n, p["base"], p["lam"], p.get("noise", 0))
        else:
            raise ValueError(f"Trend inconnu: {spec.trend_fn}")

        rows_data = [(d, pt, v) for (d, pt), v in zip(dates, vals)]
        xlsx = make_vector_xlsx(rows_data)
        result = post_file("/outstanding-vectors", spec.name, spec.desc, xlsx)
        ids[spec.name] = result["id"]
        ok(spec.name, result["id"])

    return ids


# ── Étape 3 : Amortization schedules ──────────────────────────────────────────

class ScheduleSpec(NamedTuple):
    name: str
    desc: str
    profile_fn: str      # "amort" | "nmd"
    params: dict

SCHEDULE_SPECS: list[ScheduleSpec] = [
    ScheduleSpec(
        "Schedule 20Y Linear",
        "Profil d'amortissement linéaire 20 ans (240 mois, curvature 1.5)",
        "amort",
        {"maturity": 240, "curvature": 1.5, "curvature_std": 0.15},
    ),
    ScheduleSpec(
        "Schedule 5Y Standard",
        "Profil d'amortissement 5 ans (60 mois, curvature 1.8)",
        "amort",
        {"maturity": 60, "curvature": 1.8, "curvature_std": 0.2},
    ),
    ScheduleSpec(
        "Schedule NMD Core 10Y",
        "Profil comportemental dépôts à vue — décroissance exponentielle core (lam=0.05, core=75%)",
        "nmd",
        {"lam": 0.05, "core_ratio": 0.75, "noise": 0.003},
    ),
    ScheduleSpec(
        "Schedule 7Y Corporate",
        "Profil annuité corporates 7 ans (84 mois, curvature 2.0)",
        "amort",
        {"maturity": 84, "curvature": 2.0, "curvature_std": 0.1},
    ),
    ScheduleSpec(
        "Schedule 2Y Consumer",
        "Profil linéaire crédits conso 2 ans (24 mois, curvature 1.0)",
        "amort",
        {"maturity": 24, "curvature": 1.0, "curvature_std": 0.3},
    ),
    ScheduleSpec(
        "Schedule 15Y CRE",
        "Profil amortissement immobilier commercial 15 ans (180 mois, curvature 2.0)",
        "amort",
        {"maturity": 180, "curvature": 2.0, "curvature_std": 0.1},
    ),
]


def seed_schedules() -> dict[str, str]:
    """Returns {name → id}"""
    section("3 — Amortization schedules")
    dates = all_dates()
    n = len(dates)
    ids: dict[str, str] = {}

    for spec in SCHEDULE_SPECS:
        p = spec.params
        rows_data: list[tuple[str, str, list[float]]] = []
        for i, (date, pt) in enumerate(dates):
            if spec.profile_fn == "amort":
                # Légère variation de curvature dans le temps (porte de stock)
                c_std = p.get("curvature_std", 0.0)
                c = max(0.1, p["curvature"] + (random.gauss(0, c_std) if c_std else 0.0))
                buckets = amort_at_tenors(p["maturity"], c)
            elif spec.profile_fn == "nmd":
                noise = p.get("noise", 0.0)
                buckets = nmd_at_tenors(p["lam"], p["core_ratio"])
                # petite perturbation
                buckets = [max(0.0, b + random.gauss(0, noise)) for b in buckets]
            else:
                raise ValueError(f"Profile inconnu: {spec.profile_fn}")
            rows_data.append((date, pt, buckets))

        xlsx = make_schedule_xlsx(rows_data, TENORS)
        result = post_file("/amort-schedules", spec.name, spec.desc, xlsx)
        ids[spec.name] = result["id"]
        ok(spec.name, result["id"])

    return ids


# ── Étape 4 : Portfolios + pairs ───────────────────────────────────────────────

def seed_portfolios(vector_ids: dict[str, str],
                    schedule_ids: dict[str, str]) -> dict[str, dict]:
    """
    Returns {portfolio_name → {"id": ..., "pairs": {label → pair_id}}}
    """
    section("4 — Portfolios + pairs")

    portfolio_defs = [
        {
            "name": "Retail Mortgages Portfolio",
            "description": "Portefeuille hypothécaire résidentiel — horizon 20 ans",
            "vectors":   ["Retail Mortgages 20Y"],
            "schedules": ["Schedule 20Y Linear"],
            "pairs": [
                {"vector": "Retail Mortgages 20Y", "schedule": "Schedule 20Y Linear",
                 "label": "Mortgages 20Y — Base"},
            ],
        },
        {
            "name": "SME & Corporate Portfolio",
            "description": "Prêts PME et entreprises, horizons 5-7 ans",
            "vectors":   ["SME Loans 5Y", "Corporate Loans 7Y"],
            "schedules": ["Schedule 5Y Standard", "Schedule 7Y Corporate"],
            "pairs": [
                {"vector": "SME Loans 5Y",      "schedule": "Schedule 5Y Standard",
                 "label": "SME 5Y — Base"},
                {"vector": "Corporate Loans 7Y", "schedule": "Schedule 7Y Corporate",
                 "label": "Corporate 7Y — Base"},
            ],
        },
        {
            "name": "NMD Deposits Portfolio",
            "description": "Dépôts à vue retail — modèle comportemental NMD",
            "vectors":   ["NMD Demand Deposits Core"],
            "schedules": ["Schedule NMD Core 10Y"],
            "pairs": [
                {"vector": "NMD Demand Deposits Core", "schedule": "Schedule NMD Core 10Y",
                 "label": "NMD Deposits — Core"},
            ],
        },
        {
            "name": "CRE Portfolio",
            "description": "Immobilier commercial Paris — horizon 15 ans",
            "vectors":   ["Commercial Real Estate 15Y"],
            "schedules": ["Schedule 15Y CRE"],
            "pairs": [
                {"vector": "Commercial Real Estate 15Y", "schedule": "Schedule 15Y CRE",
                 "label": "CRE 15Y — Base"},
            ],
        },
        {
            "name": "Consumer Credit Portfolio",
            "description": "Crédits consommation court terme 2 ans",
            "vectors":   ["Consumer Credit 2Y"],
            "schedules": ["Schedule 2Y Consumer"],
            "pairs": [
                {"vector": "Consumer Credit 2Y", "schedule": "Schedule 2Y Consumer",
                 "label": "Consumer 2Y — Base"},
            ],
        },
    ]

    result: dict[str, dict] = {}

    for pdef in portfolio_defs:
        # Créer le portfolio
        p = post_json("/portfolios", {"name": pdef["name"], "description": pdef["description"]})
        pid = p["id"]
        ok(pdef["name"], pid)

        # Ajouter les vecteurs
        for vname in pdef["vectors"]:
            vid = vector_ids[vname]
            requests.post(f"{API}/portfolios/{pid}/vectors",
                          json={"id": vid}, timeout=10).raise_for_status()

        # Ajouter les schedules
        for sname in pdef["schedules"]:
            sid = schedule_ids[sname]
            requests.post(f"{API}/portfolios/{pid}/schedules",
                          json={"id": sid}, timeout=10).raise_for_status()

        # Créer les paires
        pairs: dict[str, str] = {}
        for pair_def in pdef["pairs"]:
            pair = post_json(f"/portfolios/{pid}/pairs", {
                "vector_id":   vector_ids[pair_def["vector"]],
                "schedule_id": schedule_ids[pair_def["schedule"]],
                "label":       pair_def["label"],
            })
            pairs[pair_def["label"]] = pair["id"]
            print(f"    pair '{pair_def['label']}' → {pair['id']}")

        result[pdef["name"]] = {"id": pid, "pairs": pairs}

    return result


# ── Étape 5 : Hypercubes ───────────────────────────────────────────────────────

def seed_hypercubes(matrix_ids: dict[str, str]) -> dict[str, dict]:
    """Returns {name → {id, combinations}}"""
    section("5 — Hypercubes")

    hc_defs = [
        {
            "name":             "Hypercube EUR — Base + Credit",
            "description":      "Taux sans risque EUR (ESTR) + Spread crédit senior + TLP",
            "start_date":       "2022-01-01",
            "end_date":         "2024-12-01",
            "proj_end_date":    "2025-12-01",
            "time_granularity": "monthly",
            "status":           "active",
            "matrix_names":     ["Base Rate EUR (ESTR)", "Credit Spread EUR",
                                  "Term Liquidity Premium EUR"],
        },
        {
            "name":             "Hypercube EUR — Full Decomposition",
            "description":      "ESTR + Spread crédit + TLP + Charge liquidité réglementaire",
            "start_date":       "2022-01-01",
            "end_date":         "2024-12-01",
            "proj_end_date":    "2025-12-01",
            "time_granularity": "monthly",
            "status":           "active",
            "matrix_names":     ["Base Rate EUR (ESTR)", "Credit Spread EUR",
                                  "Term Liquidity Premium EUR",
                                  "Charge Liquidité Réglementaire EUR"],
        },
        {
            "name":             "Hypercube USD — SOFR",
            "description":      "Taux sans risque USD (SOFR)",
            "start_date":       "2022-01-01",
            "end_date":         "2024-12-01",
            "proj_end_date":    "2025-12-01",
            "time_granularity": "monthly",
            "status":           "active",
            "matrix_names":     ["Base Rate USD (SOFR)", "Credit Spread EUR"],
        },
    ]

    result: dict[str, dict] = {}

    for hdef in hc_defs:
        mids = []
        missing = []
        for mname in hdef["matrix_names"]:
            if mname not in matrix_ids:
                missing.append(mname)
            else:
                mids.append(matrix_ids[mname])
        if missing:
            print(f"  ⚠ Skipped '{hdef['name']}': matrices manquantes: {missing}")
            continue

        hc = post_json("/hypercubes", {
            "name":             hdef["name"],
            "description":      hdef["description"],
            "start_date":       hdef["start_date"],
            "end_date":         hdef["end_date"],
            "proj_end_date":    hdef.get("proj_end_date"),
            "time_granularity": hdef["time_granularity"],
            "status":           hdef["status"],
            "matrix_ids":       mids,
        })
        hcid = hc["id"]
        ok(hdef["name"], hcid)

        # Récupérer les combinaisons générées
        combos = get_json(f"/hypercubes/{hcid}/combinations")
        print(f"    → {len(combos)} combinaison(s)")
        for i, c in enumerate(combos):
            print(f"       [{i}] risques: {c['risks_covered']}")

        result[hdef["name"]] = {"id": hcid, "combinations": combos}

    return result


# ── Étape 6 : Study units + assignments ───────────────────────────────────────

def seed_study_units(portfolio_data: dict[str, dict],
                     hc_data: dict[str, dict]) -> dict[str, str]:
    """Returns {name → id}"""
    section("6 — Study units + assignments")

    primary_hc = "Hypercube EUR — Base + Credit"
    full_hc    = "Hypercube EUR — Full Decomposition"

    if primary_hc not in hc_data:
        print("  ⚠ Hypercube primaire non disponible — skip study units")
        return {}

    hcid_primary = hc_data[primary_hc]["id"]
    combo_primary = hc_data[primary_hc]["combinations"]
    combo0_ids   = combo_primary[0]["matrix_ids"] if combo_primary else []

    hcid_full   = hc_data.get(full_hc, {}).get("id")
    combo_full  = hc_data.get(full_hc, {}).get("combinations", [])
    combo_full0 = combo_full[0]["matrix_ids"] if combo_full else []

    su_defs = [
        {
            "name":          "SU — Retail Mortgages",
            "description":   "Study unit portefeuille hypothécaire 20 ans",
            "hypercube_id":  hcid_primary,
            "portfolio":     "Retail Mortgages Portfolio",
            "start_date":    "2022-01-01",
            "granularity_rule": "none",
            "assignments": [
                {
                    "pair_label":         "Mortgages 20Y — Base",
                    "combination_ids":    combo0_ids,
                    "label":              "Mortgages — Base rate + Credit + TLP",
                    "is_existing_stock":  False,
                },
            ],
        },
        {
            "name":          "SU — SME Portfolio",
            "description":   "Study unit prêts PME 5 ans",
            "hypercube_id":  hcid_primary,
            "portfolio":     "SME & Corporate Portfolio",
            "start_date":    "2022-01-01",
            "granularity_rule": "none",
            "assignments": [
                {
                    "pair_label":         "SME 5Y — Base",
                    "combination_ids":    combo0_ids,
                    "label":              "SME — Base rate + Credit + TLP",
                    "is_existing_stock":  False,
                },
            ],
        },
        {
            "name":          "SU — Corporate Portfolio",
            "description":   "Study unit prêts corporate 7 ans",
            "hypercube_id":  hcid_primary,
            "portfolio":     "SME & Corporate Portfolio",
            "start_date":    "2022-01-01",
            "granularity_rule": "none",
            "assignments": [
                {
                    "pair_label":         "Corporate 7Y — Base",
                    "combination_ids":    combo0_ids,
                    "label":              "Corporate — Base rate + Credit + TLP",
                    "is_existing_stock":  False,
                },
            ],
        },
        {
            "name":          "SU — NMD Deposits",
            "description":   "Study unit dépôts à vue comportementaux",
            "hypercube_id":  hcid_primary,
            "portfolio":     "NMD Deposits Portfolio",
            "start_date":    "2022-01-01",
            "granularity_rule": "none",
            "assignments": [
                {
                    "pair_label":         "NMD Deposits — Core",
                    "combination_ids":    combo0_ids,
                    "label":              "NMD Core — Base rate + Credit + TLP",
                    "is_existing_stock":  False,
                },
            ],
        },
    ]
    # Optionally add CRE with full hypercube
    if hcid_full and combo_full0:
        su_defs.append({
            "name":          "SU — CRE Full Decomposition",
            "description":   "Study unit CRE avec décomposition complète ESTR+credit+TLP+CLP",
            "hypercube_id":  hcid_full,
            "portfolio":     "CRE Portfolio",
            "start_date":    "2022-01-01",
            "granularity_rule": "none",
            "assignments": [
                {
                    "pair_label":      "CRE 15Y — Base",
                    "combination_ids": combo_full0,
                    "label":           "CRE — Full decomposition",
                    "is_existing_stock": False,
                },
            ],
        })

    ids: dict[str, str] = {}

    for sudef in su_defs:
        pname = sudef["portfolio"]
        if pname not in portfolio_data:
            print(f"  ⚠ Portfolio '{pname}' non disponible — skip {sudef['name']}")
            continue

        pdata = portfolio_data[pname]

        su = post_json("/study-units", {
            "name":             sudef["name"],
            "description":      sudef.get("description"),
            "hypercube_id":     sudef["hypercube_id"],
            "portfolio_id":     pdata["id"],
            "start_date":       sudef["start_date"],
            "granularity_rule": sudef["granularity_rule"],
        })
        suid = su["id"]
        ok(sudef["name"], suid)

        for adef in sudef["assignments"]:
            pair_label = adef["pair_label"]
            pair_id = pdata["pairs"].get(pair_label)
            if not pair_id:
                print(f"    ⚠ Pair '{pair_label}' non trouvée dans '{pname}'")
                continue

            assignment = post_json(f"/study-units/{suid}/assignments", {
                "pair_id":               pair_id,
                "combination_matrix_ids": adef["combination_ids"],
                "label":                 adef.get("label"),
                "is_existing_stock":     adef.get("is_existing_stock", False),
            })
            print(f"    assignment '{adef.get('label', '')}' → {assignment['id']}")

        ids[sudef["name"]] = suid

    return ids


# ── Étape 7 : Validation study units ─────────────────────────────────────────

def validate_study_units(su_ids: dict[str, str]) -> set[str]:
    """Returns set of valid study unit IDs."""
    section("7 — Validation study units")
    valid_ids: set[str] = set()

    for name, suid in su_ids.items():
        r = requests.post(f"{API}/study-units/{suid}/validate", timeout=30)
        if not r.ok:
            print(f"  ✗ {name}: HTTP {r.status_code}")
            continue
        report = r.json()
        if report.get("is_valid"):
            valid_ids.add(suid)
            print(f"  ✓ {name}")
        else:
            checks = report.get("checks", [])
            failed = [c["message"] for c in checks if not c["passed"]]
            print(f"  ✗ {name}:")
            for msg in failed:
                print(f"      → {msg}")

    return valid_ids


# ── Étape 8 : Studies ─────────────────────────────────────────────────────────

def seed_studies(su_ids: dict[str, str], valid_su_ids: set[str]) -> dict[str, str]:
    section("8 — Studies")

    study_defs = [
        {
            "name":        "Study — Retail & SME 2022-2025",
            "description": "Analyse FTP portefeuilles retail et PME — horizon complet",
            "units": ["SU — Retail Mortgages", "SU — SME Portfolio"],
        },
        {
            "name":        "Study — Full Book 2022-2025",
            "description": "Book complet : mortgages, corporate, SME, NMD, CRE",
            "units": ["SU — Retail Mortgages", "SU — SME Portfolio",
                      "SU — Corporate Portfolio", "SU — NMD Deposits",
                      "SU — CRE Full Decomposition"],
        },
    ]

    ids: dict[str, str] = {}

    for sdef in study_defs:
        # Déterminer le statut : ready si toutes les unités présentes sont valides
        present_units = [u for u in sdef["units"] if u in su_ids]
        all_valid = all(su_ids[u] in valid_su_ids for u in present_units if u in su_ids)

        study = post_json("/studies", {
            "name":        sdef["name"],
            "description": sdef["description"],
            "status":      "ready" if (present_units and all_valid) else "draft",
        })
        stid = study["id"]
        ok(sdef["name"], stid)

        for uname in present_units:
            suid = su_ids[uname]
            requests.post(
                f"{API}/studies/{stid}/units",
                json={"study_unit_id": suid, "label": uname},
                timeout=10,
            ).raise_for_status()
            print(f"    + {uname}")

        ids[sdef["name"]] = stid

    return ids


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Seed FTP Simulator via API")
    parser.add_argument("--skip-matrices", action="store_true",
                        help="Ne pas re-générer les matrices de taux")
    parser.add_argument("--skip-study-units", action="store_true",
                        help="Arrêter après les portfolios et hypercubes")
    args = parser.parse_args()

    # Vérifier que le backend répond
    try:
        requests.get(f"{API}/risk-types", timeout=5).raise_for_status()
    except Exception as e:
        print(f"\n✗ Backend inaccessible sur {API}: {e}")
        print("  Démarrez le backend avec : cargo run --release")
        sys.exit(1)

    print(f"\nBackend OK — {API}")

    # ── 1. Matrices ──
    if not args.skip_matrices:
        seed_matrices()
    else:
        print("\n[skip matrices]")

    # Récupérer les IDs des matrices depuis l'API
    matrices = get_json("/rate-matrices")
    matrix_ids: dict[str, str] = {m["name"]: m["id"] for m in matrices}
    print(f"\n  {len(matrix_ids)} matrice(s) en base : {list(matrix_ids.keys())}")

    if not matrix_ids:
        print("  ⚠ Aucune matrice en base. Vérifiez l'étape 1.")
        sys.exit(1)

    # ── 2-3. Vectors & Schedules ──
    vector_ids   = seed_vectors()
    schedule_ids = seed_schedules()

    # ── 4. Portfolios ──
    portfolio_data = seed_portfolios(vector_ids, schedule_ids)

    # ── 5. Hypercubes ──
    hc_data = seed_hypercubes(matrix_ids)

    if args.skip_study_units:
        print("\n[stop après hypercubes]")
        _print_summary(matrix_ids, vector_ids, schedule_ids, portfolio_data, hc_data, {}, {}, {})
        return

    # ── 6. Study units ──
    su_ids = seed_study_units(portfolio_data, hc_data)

    # ── 7. Validation ──
    valid_su_ids = validate_study_units(su_ids)

    # ── 8. Studies ──
    study_ids = seed_studies(su_ids, valid_su_ids)

    _print_summary(matrix_ids, vector_ids, schedule_ids, portfolio_data,
                   hc_data, su_ids, valid_su_ids, study_ids)


def _print_summary(matrix_ids, vector_ids, schedule_ids, portfolio_data,
                   hc_data, su_ids, valid_su_ids, study_ids):
    section("RÉSUMÉ")
    print(f"  Matrices         : {len(matrix_ids)}")
    print(f"  Vectors          : {len(vector_ids)}")
    print(f"  Schedules        : {len(schedule_ids)}")
    print(f"  Portfolios       : {len(portfolio_data)}")
    print(f"  Hypercubes       : {len(hc_data)}")
    print(f"  Study units      : {len(su_ids)}  (valides : {len(valid_su_ids)})")
    print(f"  Studies          : {len(study_ids)}")
    if study_ids:
        ready = [n for n, sid in study_ids.items()]
        print(f"\n  Pour lancer une exécution, utilisez l'UI (/executions)")
        print(f"  ou : curl -X POST {API}/executions \\")
        study_id = list(study_ids.values())[0]
        print(f'       -H "Content-Type: application/json" \\')
        print(f'       -d \'{{"study_id":"{study_id}"}}\' ')
    print()


if __name__ == "__main__":
    main()
